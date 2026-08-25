from __future__ import annotations

import json
import importlib.util
import re
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Dict, Iterable, List, Mapping, Sequence, Tuple

from ..history.database import Database
from ..history.metrics import summarize_usage


class ReportingDependencyError(RuntimeError):
    pass


def generate_workbook(db: Database, output_dir: Path, kear_id: str, logical_name: str,
                      application_labels: Sequence[str], environment: str,
                      lookback_days: int, as_of: date) -> Path:
    if not kear_id.strip():
        raise ValueError("kear_id must not be empty")
    if not logical_name.strip() or not application_labels or not environment.strip():
        raise ValueError("logical name, application labels, and environment are required")
    if importlib.util.find_spec("openpyxl") is None:
        raise ReportingDependencyError("openpyxl is required for report generation; install the approved offline package")
    from openpyxl import Workbook
    kear = kear_id.lower()
    cutoff = (as_of - timedelta(days=lookback_days)).isoformat()
    with db.connect() as connection:
        rule_rows = [dict(row) for row in connection.execute("SELECT * FROM rules ORDER BY ruleset_name,rule_href")]
        selected = [row for row in rule_rows if _rule_matches(row, application_labels, environment)]
        usage_by_rule: Dict[str, List[Mapping[str, object]]] = defaultdict(list)
        for row in connection.execute("SELECT * FROM usage_windows WHERE window_end>? AND window_start<? ORDER BY window_start", (cutoff, as_of.isoformat())):
            usage_by_rule[row["rule_href"]].append(dict(row))
        workloads = [dict(row) for row in connection.execute("SELECT * FROM workloads")]
        quality = [dict(row) for row in connection.execute("SELECT * FROM data_quality ORDER BY category,object_id")]
    raw_rows, expanded_rows, usage_rows = [], [], []
    for rule in selected:
        raw = json.loads(rule["raw_json"])
        modules = [label for label in application_labels if _contains_app(raw, label)]
        metrics = summarize_usage(usage_by_rule[rule["rule_href"]], as_of, lookback_days)
        base = {
            "KEAR ID": kear, "Logical Application": logical_name, "Module": "\n".join(modules),
            "Environment": environment, "Ruleset": rule["ruleset_name"], "Rule Href": rule["rule_href"],
            "Ruleset Enabled": _display_bool(rule["ruleset_enabled"]), "Rule Enabled": _display_bool(rule["rule_enabled"]),
            "Rule Type": rule["rule_type"], "Source": rule["source_text"], "Destination": rule["destination_text"],
            "Service Name / Definition": rule["services"], "Description": rule["rule_description"],
            "Hit Status": metrics["hit_status"], "Total Flows": metrics["total_flows"],
            "First Hit Window": _window(metrics["first_hit_window_start"], metrics["first_hit_window_end"]),
            "Last Hit Window": _window(metrics["last_hit_window_start"], metrics["last_hit_window_end"]),
            "Days Since Last Hit": metrics["days_since_last_hit"], "Coverage %": metrics["coverage_percent"],
        }
        raw_rows.append(base)
        expanded = dict(base)
        expanded["Expanded Sources"] = _expand_side(raw, "src", workloads, environment)
        expanded["Expanded Destinations"] = _expand_side(raw, "dst", workloads, environment)
        expanded_rows.append(expanded)
        for usage in usage_by_rule[rule["rule_href"]]:
            usage_rows.append({"KEAR ID": kear, "Rule Href": rule["rule_href"], "Window Start": usage["window_start"],
                               "Window End": usage["window_end"], "Status": usage["status"], "Flows": usage["flows"],
                               "Port Breakdown Complete": _display_bool(usage["port_breakdown_complete"]),
                               "Omitted Port Details": usage["port_details_omitted_count"]})
    workbook = Workbook(); workbook.remove(workbook.active)
    presentation = [{"Field": key, "Value": value} for key, value in (
        ("KEAR ID", kear), ("Logical Application", logical_name), ("Application Labels", "\n".join(application_labels)),
        ("Environment", environment), ("As Of", as_of.isoformat()), ("Lookback Days", lookback_days),
        ("Rules", len(selected)), ("Generated At UTC", datetime.now(timezone.utc).isoformat()))]
    _sheet(workbook, "Presentation", presentation)
    _sheet(workbook, "Raw Rules", raw_rows)
    _sheet(workbook, "Expanded Rules", expanded_rows)
    _sheet(workbook, "Rule Usage", usage_rows)
    quality_rows = [{"KEAR ID": kear, **row} for row in quality]
    _sheet(workbook, "Data Quality", quality_rows)
    output_dir.mkdir(parents=True, exist_ok=True)
    safe_env = re.sub(r"[^A-Za-z0-9_.-]+", "_", environment)
    target = output_dir / f"rules_recertify_{kear}_{safe_env}_{as_of.strftime('%Y%m%d')}.xlsx"
    temporary = target.with_suffix(".xlsx.tmp")
    workbook.save(temporary); temporary.replace(target)
    return target


def _rule_matches(rule: Mapping[str, object], labels: Sequence[str], environment: str) -> bool:
    raw = json.loads(str(rule["raw_json"]))
    scope = str(raw.get("ruleset_scope", ""))
    env_values = [part[4:] for part in scope.split(";") if part.lower().startswith("env:")]
    env_match = not env_values or any(value.upper() in {"NULL", environment.upper()} for value in env_values)
    return env_match and any(_contains_app(raw, label) for label in labels)


def _contains_app(raw: Mapping[str, object], label: str) -> bool:
    needle = f"app:{label}".lower()
    fields = ("ruleset_scope", "src_labels", "dst_labels", "src_labels_exclusions", "dst_labels_exclusions")
    for field in fields:
        tokens = {token.strip().lower() for token in str(raw.get(field, "")).split(";")}
        if needle in tokens:
            return True
    return False


def _expand_side(raw: Mapping[str, object], side: str, workloads: Sequence[Mapping[str, object]], environment: str) -> str:
    values: List[str] = []
    if str(raw.get(f"{side}_all_workloads", "")).lower() == "true": values.append("All Workloads")
    ip_lists = str(raw.get(f"{side}_iplists", ""));
    if ip_lists: values.extend(f"IP List: {item.strip()}" for item in ip_lists.split(";") if item.strip())
    explicit = str(raw.get(f"{side}_workloads", ""));
    if explicit: values.extend(f"Workload: {item.strip()}" for item in explicit.split(";") if item.strip())
    labels = str(raw.get(f"{side}_labels", ""))
    if labels:
        wanted = {part.strip() for part in labels.split(";") if part.strip()}
        for workload in workloads:
            if workload.get("env") != environment: continue
            tags = {f"app:{workload.get('app','')}", f"env:{workload.get('env','')}", f"loc:{workload.get('loc','')}", f"role:{workload.get('role','')}"}
            if wanted.issubset(tags):
                for address in json.loads(str(workload["addresses_json"])):
                    values.append(f"{workload.get('hostname') or workload.get('name')} ({address})")
    if "Any (0.0.0.0/0 and ::/0)" in ip_lists: values.extend(["0.0.0.0/0", "::/0"])
    return "\n".join(dict.fromkeys(values))


def _sheet(workbook: object, name: str, rows: List[Mapping[str, object]]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    sheet = workbook.create_sheet(name)
    headers = list(rows[0]) if rows else ["KEAR ID", "Message"]
    sheet.append(headers)
    for row in rows: sheet.append([row.get(header) for header in headers])
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78"); cell.font = Font(color="FFFFFF", bold=True)
    sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        letter = column[0].column_letter
        width = min(60, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
        sheet.column_dimensions[letter].width = width
        for cell in column: cell.alignment = Alignment(vertical="top", wrap_text=True)


def _display_bool(value: object) -> str:
    return "TRUE" if value == 1 else "FALSE" if value == 0 else ""


def _window(start: object, end: object) -> str:
    return f"{start} / {end}" if start and end else ""
